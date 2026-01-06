"""
Excel Reader - Read previous GSTR-1 Excel files for continuity checking.

Reads files with name starting with "GSTR1-Portal-" and extracts
document series information from the "Doc Issued" sheet.
"""

import os
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass


@dataclass
class PreviousSeriesInfo:
    """Information about a series from previous GSTR-1."""
    doc_type: str
    from_invoice: str
    to_invoice: str
    total: int
    cancelled: int
    tax_period: str = ""  # Tax period in MMYYYY format (e.g., "042024")

    # Extracted pattern info
    prefix: str = ""  # Everything before the sequence number
    suffix: str = ""  # Everything after the sequence number
    end_sequence: int = 0  # The ending sequence number


def read_gstr1_excel(filepath: str) -> Tuple[List[PreviousSeriesInfo], str]:
    """
    Read a GSTR-1 Portal Excel file and extract series information.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        Tuple of (list of PreviousSeriesInfo, message)
    """
    # Validate filename
    filename = os.path.basename(filepath)
    if not filename.startswith("GSTR1-Portal-"):
        return [], f"File must start with 'GSTR1-Portal-'. Got: {filename}"
    
    # Check extension
    if not filepath.lower().endswith('.xlsx'):
        return [], "File must be an Excel file (.xlsx)"
    
    # Try to import openpyxl
    try:
        import openpyxl
    except ImportError:
        return [], "Excel support requires 'openpyxl'. Install with: pip install openpyxl"
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return [], f"Could not open Excel file: {str(e)}"
    
    # Find the DocIssued sheet
    sheet = None
    sheet_names_to_try = ["DocIssued", "docissued", "DOCISSUED", "b2b", "B2B"]
    
    for name in sheet_names_to_try:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    
    # If not found by name, try first sheet
    if sheet is None:
        sheet = wb.active
    
    if sheet is None:
        return [], "Could not find document data in the Excel file"
    
    # Find header row and column indices
    # Expected columns: Company GSTIN, Tax Period, Doc Type, From, To, Total, Cancelled, GSTR-1A
    headers = {}
    header_row = None

    for row_idx, row in enumerate(sheet.iter_rows(max_row=15, values_only=True), 1):
        row_lower = [str(cell).lower().strip() if cell else "" for cell in row]

        # Look for key columns
        if any('doc type' in cell or 'nature' in cell for cell in row_lower):
            header_row = row_idx
            for col_idx, cell in enumerate(row_lower):
                if 'doc type' in cell or 'nature' in cell:
                    headers['doc_type'] = col_idx
                elif cell == 'from':
                    headers['from'] = col_idx
                elif cell == 'to':
                    headers['to'] = col_idx
                elif cell == 'total':
                    headers['total'] = col_idx
                elif cell == 'cancelled':
                    headers['cancelled'] = col_idx
                elif 'tax period' in cell or 'period' in cell:
                    headers['tax_period'] = col_idx
            break
    
    if header_row is None or 'doc_type' not in headers or 'from' not in headers or 'to' not in headers:
        return [], "Could not find required columns (Doc Type, From, To) in the Excel file"
    
    # Read data rows
    series_list = []

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        doc_type = str(row[headers['doc_type']]).strip() if row[headers['doc_type']] else ""
        from_inv = str(row[headers['from']]).strip() if row[headers['from']] else ""
        to_inv = str(row[headers['to']]).strip() if row[headers['to']] else ""

        # Skip empty rows
        if not doc_type or not from_inv or not to_inv:
            continue

        # Skip header-like rows that might be repeated
        if 'doc type' in doc_type.lower() or 'from' in doc_type.lower():
            continue

        total = 0
        cancelled = 0
        tax_period = ""

        if 'total' in headers and row[headers['total']]:
            try:
                total = int(row[headers['total']])
            except (ValueError, TypeError):
                pass

        if 'cancelled' in headers and row[headers['cancelled']]:
            try:
                cancelled = int(row[headers['cancelled']])
            except (ValueError, TypeError):
                pass

        if 'tax_period' in headers and row[headers['tax_period']]:
            tax_period_raw = str(row[headers['tax_period']]).strip()
            tax_period = _parse_tax_period_from_excel(tax_period_raw)

        info = PreviousSeriesInfo(
            doc_type=doc_type,
            from_invoice=from_inv,
            to_invoice=to_inv,
            total=total,
            cancelled=cancelled,
            tax_period=tax_period
        )

        # Extract prefix/suffix from the 'to' invoice
        _extract_pattern_info(info)

        series_list.append(info)
    
    wb.close()
    
    if not series_list:
        return [], "No document series found in the Excel file"
    
    return series_list, f"Found {len(series_list)} series from previous GSTR-1"


def _parse_tax_period_from_excel(raw_value: str) -> str:
    """
    Parse tax period from Excel cell value to MMYYYY format.

    Handles formats like:
    - "042024" (already in MMYYYY format)
    - "04/2024"
    - "Apr 2024"
    - "April 2024"
    - etc.

    Args:
        raw_value: Raw tax period value from Excel

    Returns:
        Tax period in MMYYYY format, or empty string if cannot parse
    """
    import re

    raw_value = raw_value.strip()

    # Already in MMYYYY format
    if re.match(r'^\d{6}$', raw_value):
        return raw_value

    # MM/YYYY or MM-YYYY format
    match = re.match(r'^(\d{1,2})[-/](\d{4})$', raw_value)
    if match:
        month = int(match.group(1))
        year = match.group(2)
        if 1 <= month <= 12:
            return f"{month:02d}{year}"

    # Month name YYYY format (e.g., "Apr 2024", "April 2024")
    month_names = {
        "jan": "01", "january": "01",
        "feb": "02", "february": "02",
        "mar": "03", "march": "03",
        "apr": "04", "april": "04",
        "may": "05",
        "jun": "06", "june": "06",
        "jul": "07", "july": "07",
        "aug": "08", "august": "08",
        "sep": "09", "september": "09",
        "oct": "10", "october": "10",
        "nov": "11", "november": "11",
        "dec": "12", "december": "12"
    }

    parts = raw_value.split()
    if len(parts) == 2:
        month_str, year_str = parts
        month_code = month_names.get(month_str.lower())
        if month_code and re.match(r'^\d{4}$', year_str):
            return f"{month_code}{year_str}"

    return ""


def _extract_pattern_info(info: PreviousSeriesInfo) -> None:
    """
    Extract prefix, suffix, and end sequence from the 'to' invoice number.
    
    For example: "GST/24-25/0500" -> prefix="GST/24-25/", suffix="", end_sequence=500
    """
    import re
    
    invoice = info.to_invoice
    
    # Find all numeric sequences
    matches = list(re.finditer(r'\d+', invoice))
    
    if not matches:
        return
    
    # Usually the last numeric sequence is the invoice number
    # But check if there's one that looks like a sequence (not year)
    best_match = None
    
    for match in matches:
        num_str = match.group()
        num_val = int(num_str)
        
        # Skip likely years (19xx, 20xx) or short year codes (23, 24, 25)
        if 1900 <= num_val <= 2100:
            continue
        if len(num_str) == 2 and 0 <= num_val <= 50:
            # Could be year code like 24, 25
            continue
        
        best_match = match
    
    # If no good match, use the last number
    if best_match is None and matches:
        best_match = matches[-1]
    
    if best_match:
        info.prefix = invoice[:best_match.start()]
        info.suffix = invoice[best_match.end():]
        info.end_sequence = int(best_match.group())


def match_series(
    previous: List[PreviousSeriesInfo],
    current_series_key: str,
    current_start_invoice: str
) -> Optional[Tuple[PreviousSeriesInfo, bool, str]]:
    """
    Find matching previous series for a current series.
    
    Args:
        previous: List of previous series info
        current_series_key: The series key from current month
        current_start_invoice: The starting invoice of current month
        
    Returns:
        Tuple of (matched_previous_series, is_continuous, message)
        Returns None if no match found
    """
    import re
    
    # Extract prefix/suffix from current start invoice
    matches = list(re.finditer(r'\d+', current_start_invoice))
    
    if not matches:
        return None
    
    # Find the sequence number (similar logic as _extract_pattern_info)
    best_match = None
    for match in matches:
        num_str = match.group()
        num_val = int(num_str)
        if 1900 <= num_val <= 2100:
            continue
        if len(num_str) == 2 and 0 <= num_val <= 50:
            continue
        best_match = match
    
    if best_match is None and matches:
        best_match = matches[-1]
    
    if not best_match:
        return None
    
    current_prefix = current_start_invoice[:best_match.start()]
    current_suffix = current_start_invoice[best_match.end():]
    current_start_seq = int(best_match.group())
    
    # Find matching previous series by prefix/suffix
    for prev in previous:
        # Match by prefix and suffix (ignoring minor differences like year)
        # Exact match first
        if prev.prefix == current_prefix and prev.suffix == current_suffix:
            expected_next = prev.end_sequence + 1
            is_continuous = (current_start_seq == expected_next)
            
            if is_continuous:
                msg = f"✓ Continuity OK: Previous ended at {prev.to_invoice}, current starts at {current_start_invoice}"
            else:
                gap = current_start_seq - prev.end_sequence - 1
                if gap > 0:
                    msg = f"⚠️ Gap detected: Previous ended at {prev.to_invoice} (seq {prev.end_sequence}), current starts at {current_start_invoice} (seq {current_start_seq}). Missing {gap} invoice(s): {prev.end_sequence + 1} to {current_start_seq - 1}"
                else:
                    msg = f"⚠️ Overlap: Previous ended at {prev.end_sequence}, current starts at {current_start_seq}"
            
            return (prev, is_continuous, msg)
    
    # Try fuzzy match (prefix similarity)
    for prev in previous:
        # Check if prefixes are similar (e.g., year change: 24-25 vs 25-26)
        if _prefixes_similar(prev.prefix, current_prefix) and prev.suffix == current_suffix:
            # This might be a year change
            expected_next = 1  # New year typically starts at 1
            is_continuous = True  # We can't really tell, assume OK for new year
            msg = f"ℹ️ New series year detected: Previous {prev.to_invoice}, current {current_start_invoice}"
            return (prev, is_continuous, msg)
    
    return None


def _prefixes_similar(prefix1: str, prefix2: str) -> bool:
    """Check if two prefixes are similar (might differ only in year)."""
    import re
    
    # Remove year-like patterns and compare
    def remove_years(s):
        # Remove patterns like 24-25, 2024-25, 23-24, etc.
        s = re.sub(r'\d{2,4}[-/]\d{2,4}', '', s)
        s = re.sub(r'\d{4}', '', s)
        return s.strip()
    
    return remove_years(prefix1) == remove_years(prefix2)
